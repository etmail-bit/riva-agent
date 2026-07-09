#!/usr/bin/env python3
"""銷售統計/明細匯入程式：把 data/raw 下的品項銷售報表讀進 raw_product_sales_monthly。

A 店（銷售統計/銷售明細）：門市名稱 / 項目名稱 / 銷售量，整月一列，沒有金額跟時段。
B 店（銷售明細）：門市名稱 / 項目名稱 / 時段 / 數量 / 金額，一個品項一個月會拆成多列（逐時段）。

兩店資料顆粒度不同：算「這個品項這個月賣幾杯」都要 SUM(quantity) GROUP BY store_id,
year_month, product_name，不能直接比較單一列的 quantity 大小。

用法：
    source .venv/bin/activate
    python3 scripts/import_product_sales.py
"""
import glob
import re
import sqlite3
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "db" / "riva_agent.db"
ENV_PATH = ROOT / ".env"
RAW_GLOBS = [
    str(ROOT / "data" / "raw" / "*銷售統計.xlsx"),
    str(ROOT / "data" / "raw" / "*銷售明細.xlsx"),
]

COLUMN_ALIASES = {
    "store_name": ["門市名稱"],
    "product_name": ["項目名稱"],
    "quantity": ["銷售量", "數量"],
    "hour_slot": ["時段"],
    "amount": ["金額"],
}


def load_store_map():
    store_map = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip()
        if key.startswith("STORE_") and key.endswith("_REAL_NAME") and value:
            store_id = key[len("STORE_"):-len("_REAL_NAME")]
            store_map[value] = store_id
    return store_map


def extract_year_month(filename):
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", filename)
    if not match:
        raise ValueError(f"{filename}：檔名裡找不到 YYYYMM，請確認檔名格式")
    return f"{match.group(1)}-{match.group(2)}"


def build_column_index(header):
    stripped_header = [str(h).strip() if h is not None else h for h in header]
    col = {}
    for unified, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in stripped_header:
                col[unified] = stripped_header.index(alias)
                break
    return col


def import_file(path, store_map, conn):
    year_month = extract_year_month(path.name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    col = build_column_index(header)

    required = {"store_name", "product_name", "quantity"}
    missing = required - col.keys()
    if missing:
        raise ValueError(f"{path.name}：缺少必要欄位 {missing}，欄位對照表可能需要更新")

    inserted = 0
    for row in data_rows:
        if row[col["product_name"]] is None:
            continue
        raw_store_name = row[col["store_name"]]
        store_id = store_map.get(raw_store_name)
        if store_id is None:
            raise ValueError(
                f"{path.name}：門市名稱在 .env 對照表裡找不到（不印出真實店名，避免洩漏），"
                "請確認 STORE_A_REAL_NAME / STORE_B_REAL_NAME"
            )

        conn.execute(
            """
            INSERT INTO raw_product_sales_monthly
                (store_id, year_month, product_name, quantity, hour_slot, amount, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                store_id,
                year_month,
                row[col["product_name"]],
                row[col["quantity"]],
                row[col["hour_slot"]] if "hour_slot" in col else None,
                row[col["amount"]] if "amount" in col else None,
                path.name,
            ),
        )
        inserted += 1
    return inserted, year_month


def main():
    store_map = load_store_map()
    if not store_map:
        raise SystemExit(".env 裡的 STORE_A_REAL_NAME / STORE_B_REAL_NAME 都是空的，請先填店名對照")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    files = sorted({Path(p) for pattern in RAW_GLOBS for p in glob.glob(pattern)})
    if not files:
        raise SystemExit("找不到符合『*銷售統計.xlsx』或『*銷售明細.xlsx』的檔案")

    # 重跑時先清掉這批檔案涵蓋到的 (store_id, year_month)，避免重複累加
    file_periods = {(path, extract_year_month(path.name)) for path in files}
    for path, year_month in file_periods:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col = build_column_index(header)
        first_row = next(
            (r for r in ws.iter_rows(min_row=2, values_only=True) if r[col["store_name"]] is not None),
            None,
        )
        if first_row is None:
            continue
        store_id = store_map.get(first_row[col["store_name"]])
        if store_id:
            conn.execute(
                "DELETE FROM raw_product_sales_monthly WHERE store_id = ? AND year_month = ?",
                (store_id, year_month),
            )

    total = 0
    for path in files:
        count, year_month = import_file(path, store_map, conn)
        print(f"{path.name} ({year_month}): 匯入 {count} 筆")
        total += count

    conn.commit()
    conn.close()
    print(f"完成，總計匯入 {total} 筆到 raw_product_sales_monthly")


if __name__ == "__main__":
    main()
