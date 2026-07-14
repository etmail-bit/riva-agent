#!/usr/bin/env python3
"""產生給使用者自助填時數算薪水的 Excel 試算表（reports/勞健保薪資試算表.xlsx，本機限定，
不進版控）。跟 scripts/calculate_payroll.py 是同一套官方級距費率邏輯的「離線/手動輸入版」——
calculate_payroll.py 吃 raw_staffing_actual 自動算，這支是給使用者自己每月手動輸入時數用，
不需要開這個系統、不需要謄打排班照片。

級距/費率數字直接從 config/insurance_rates_2026.json 讀，不手key，避免跟系統主要計算邏輯
的數字兜不起來（2026-07-14 建立，來源見該檔案的 _說明 欄位與 scripts/calculate_payroll.py
模組說明——官方三個資料點已核對過，職災保險費率沒有官方逐點金額表可核對）。

用法：
    source .venv/bin/activate
    python3 -m scripts.build_payroll_excel
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
INSURANCE_CONFIG_PATH = ROOT / "config" / "insurance_rates_2026.json"
STAFFING_CONFIG_PATH = ROOT / "config" / "staffing_rules.json"
OUT_PATH = ROOT / "reports" / "勞健保薪資試算表.xlsx"

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
GREEN = Font(name=FONT_NAME, color="008000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_configs():
    insurance = json.loads(INSURANCE_CONFIG_PATH.read_text(encoding="utf-8"))
    staffing = json.loads(STAFFING_CONFIG_PATH.read_text(encoding="utf-8"))
    return insurance, staffing


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_bracket_sheet(wb, sheet_name, brackets, title):
    """floor(下限,含) / ceiling(上限,顯示用) / insured_amount(月投保金額)，
    VLOOKUP 用「下限」欄位做約略比對(TRUE)，所以下限這欄一定要由小到大排序。"""
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = BOLD
    ws.append(["下限（元，含）", "上限（元，含）", "月投保金額（元）"])
    style_header_row(ws, 2, 3)
    floor = 0
    for b in brackets:
        ceiling = b["wage_ceiling"]
        ws.append([floor + 1 if floor else 1, ceiling if ceiling is not None else "以上", b["insured_amount"]])
        floor = ceiling if ceiling is not None else floor
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.font = BLACK
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
    autosize(ws, {"A": 16, "B": 16, "C": 16})
    ws.freeze_panes = "A3"
    return ws, ws.max_row


def main():
    insurance, staffing = load_configs()
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- 說明 ----------
    ws = wb.create_sheet("說明")
    ws.column_dimensions["A"].width = 100
    lines = [
        ("勞健保薪資試算表", BOLD),
        ("", None),
        ("怎麼用：", BOLD),
        ("1. 第一次使用先去「員工資料」分頁，確認每位員工的代碼／角色／底薪或時薪是最新的（藍色格才能改）。", None),
        ("2. 之後每個月只要去「薪資試算」分頁，藍色格填每位員工這個月的「上班時數」「加班時數」就好，其他都是公式自動算。", None),
        ("3. 想試算新月份，複製整張「薪資試算」分頁重新命名（例如「薪資試算_2026-08」），保留公式格式最方便。", None),
        ("", None),
        ("公司總負擔成本 = 底薪/工資 ＋ 加班費 ＋ 勞保／職災保險／健保／勞退四項雇主負擔。", None),
        ("不含員工自付額（表裡有列出來當參考，但沒有加進公司總成本）。", None),
        ("", None),
        ("重要提醒：", BOLD),
        (f"・資料來源：民國115年(2026年)勞保/職災/健保/勞退官方級距費率，"
         f"生效日 {insurance['effective_date']}，逐年會變動，明年要記得換新的級距表，"
         "不要沿用這份舊的（勞動部勞工保險局／衛福部中央健康保險署官網每年底會公告新年度級距表）。", None),
        ("・這裡算出來的金額是公式算的，跟勞保局/健保署官方金額表可能有 ±1 元等級的捨入落差"
         "（官方逐級距金額表本身的捨入規則沒有完全公開），適合拿來抓大概的人事成本，"
         "不是拿去申報用的正式金額。", None),
        (f"・職業災害保險費率目前用政府公告「{insurance['occupational_injury_insurance']['industry']}」"
         "大類的預設值（見「參數設定」分頁），還沒拿公司實際的勞保局繳款單核對過——"
         "繳款單上會直接印出核定費率，之後拿到了直接改「參數設定」分頁那兩個數字就好，不用改公式。", None),
        ("・健保「公司負擔」那欄，依法要內含「全國平均眷屬人數」（見「參數設定」分頁的乘數），"
         "這是公司單方面的法定負擔，不是這位員工本人有沒有真的申報眷屬的問題，也不會轉嫁給員工——"
         "員工自付那欄只算本人。", None),
        ("・如果上班時數是從照片/PDF謄打的排班表抓出來的，短班兼職的時數謄打容易出錯，"
         "建議先人工核對過一次再拿來算薪水，不要照單全收。", None),
        ("", None),
        ("跟 Riva-agent 系統的關係：", BOLD),
        ("這份表跟 scripts/calculate_payroll.py 是同一套官方級距費率邏輯，"
         "差別只在這份表是「手動輸入時數」，系統那支是「自動讀真實排班表算」。"
         "兩邊的級距/費率數字本來就是同一個來源（config/insurance_rates_2026.json），"
         "理論上同樣的時數應該算出一樣的金額。", None),
    ]
    r = 1
    for text, font in lines:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font or Font(name=FONT_NAME)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text and not font:
            ws.row_dimensions[r].height = 30
        r += 1

    # ---------- 參數設定 ----------
    ws = wb.create_sheet("參數設定")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46
    ws.append(["項目", "數值", "說明"])
    style_header_row(ws, 1, 3)

    wages = staffing["wages"]
    labor = insurance["labor_insurance"]
    occ = insurance["occupational_injury_insurance"]
    health = insurance["health_insurance"]
    pension = insurance["pension"]

    param_rows = [
        ("經理固定月薪", wages["manager_monthly"], "角色=經理 的底薪，員工資料分頁預設引用這裡"),
        ("正職固定月薪", wages["staff_monthly"], "角色=正職 的底薪（員工資料分頁可個別覆寫）"),
        ("加班時薪", wages["overtime_hourly"], "經理/正職超過「每日正常工時」的部分，以此時薪計"),
        ("每日正常工時", wages["daily_regular_hours"], "超過這個時數才算加班（本表暫不逐日判斷，直接輸入本月加班總時數）"),
        ("兼職預設時薪", wages["part_time_hourly"], "員工資料分頁可個別覆寫"),
        ("勞保_普通事故費率", labor["ordinary_rate"], "勞保投保薪資分級表(115年1月1日起適用)"),
        ("勞保_就業保險費率", labor["employment_insurance_rate"], "同上，就業保險費率"),
        ("勞保_員工比例", labor["split"]["employee"], ""),
        ("勞保_公司比例", labor["split"]["employer"], ""),
        ("職災_行業別費率", occ["industry_rate"], f"預設「{occ['industry']}」大類，有繳款單請改這裡"),
        ("職災_上下班費率", occ["commute_rate"], "全國統一費率"),
        ("健保_費率", health["rate"], "全民健康保險費率"),
        ("健保_員工比例", health["split"]["employee"], "只算本人，不含眷屬"),
        ("健保_公司比例", health["split"]["employer"], ""),
        ("健保_公司眷屬平均乘數", 1 + health["employer_avg_dependents_multiplier"],
         f"公司負擔=健保基數×公司比例×此乘數（法定內含全國平均眷屬{health['employer_avg_dependents_multiplier']}人，1+{health['employer_avg_dependents_multiplier']}={1 + health['employer_avg_dependents_multiplier']}）"),
        ("勞退_費率", pension["rate"], "雇主強制提繳，全額公司負擔"),
    ]
    for name, val, note in param_rows:
        ws.append([name, val, note])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        row[0].font = BLACK
        row[1].font = BLUE
        row[1].fill = ASSUMPTION_FILL
        row[2].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
        for c in row:
            c.border = BORDER
    ws.freeze_panes = "A2"
    PARAM = "參數設定"

    def p(name):
        row = next(i for i, (n, _, _) in enumerate(param_rows, start=2) if n == name)
        return f"'{PARAM}'!$B${row}"

    # ---------- 級距表（4張） ----------
    # 勞保／職災保險的級距表要用「pension 細分級距表」往下延伸到45,800/72,800元封頂的版本
    # （不能只用 labor/occ 原本的11/21級主表，那個只從29,500元起算）——兼職員工月收入常常
    # 低於29,500元，Excel公式對兼職是直接查這張表（不套 MAX(x,29500)），如果表本身沒有低收入
    # 細分級距，會查到表上第一格（29,500元）而不是官方規定的細分級距金額，等於錯誤墊高兼職的
    # 保費。這個bug是2026-07-14建好表後，拿已知正確答案交叉比對才抓到的，抓到後才修。
    labor_extended = [b for b in pension["brackets"] if b["insured_amount"] <= 45800]
    occ_extended = [b for b in pension["brackets"] if b["insured_amount"] <= 72800]
    _, LAB_LAST = build_bracket_sheet(wb, "勞保級距表", labor_extended,
                         "勞工保險投保薪資分級表（115年1月1日起適用，最高45,800元；"
                         "29,500元以下部分沿用勞退細分級距表，供兼職員工低收入查表用）")
    _, OCC_LAST = build_bracket_sheet(wb, "職災保險級距表", occ_extended,
                         "勞工職業災害保險投保薪資分級表（最高72,800元；"
                         "29,500元以下部分沿用勞退細分級距表，供兼職員工低收入查表用）")
    _, HEA_LAST = build_bracket_sheet(wb, "健保級距表", health["brackets"], "全民健康保險投保金額分級表（115.01.01生效，最高313,000元）")
    _, PEN_LAST = build_bracket_sheet(wb, "勞退級距表", pension["brackets"], "勞工退休金月提繳分級表（115年1月1日生效，含低於基本工資的細分級距）")

    # ---------- 員工資料 ----------
    ws = wb.create_sheet("員工資料")
    ws.append(["員工代碼", "角色", "底薪（經理/正職適用）", "時薪（兼職適用）"])
    style_header_row(ws, 1, 4)
    roles = {k: v for k, v in staffing["employee_roles"].items() if not k.startswith("_")}
    role_label = {"manager": "經理", "staff": "正職", "part_time": "兼職"}
    for code, role in sorted(roles.items()):
        label = role_label.get(role, role)
        base = "=$B$2" if label == "經理" else ("=$B$3" if label == "正職" else "")
        hourly = "=$B$6" if label == "兼職" else ""
        ws.append([code, label, base, hourly])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        row[0].font = BLUE
        row[1].font = BLUE
        row[2].font = GREEN if isinstance(row[2].value, str) and row[2].value.startswith("=") else BLUE
        row[3].font = GREEN if isinstance(row[3].value, str) and row[3].value.startswith("=") else BLUE
        for c in row:
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
    # 底薪/時薪公式改成指到參數設定分頁對應格，比純文字 "=$B$2" 更正確（跨分頁引用）
    for i, r in enumerate(range(2, ws.max_row + 1)):
        label = ws.cell(row=r, column=2).value
        if label == "經理":
            ws.cell(row=r, column=3, value=f"={p('經理固定月薪')}")
        elif label == "正職":
            ws.cell(row=r, column=3, value=f"={p('正職固定月薪')}")
        elif label == "兼職":
            ws.cell(row=r, column=4, value=f"={p('兼職預設時薪')}")
    autosize(ws, {"A": 12, "B": 12, "C": 22, "D": 18})
    dv_role = DataValidation(type="list", formula1='"經理,正職,兼職"', allow_blank=False)
    ws.add_data_validation(dv_role)
    dv_role.add(f"B2:B{ws.max_row}")
    ws.freeze_panes = "A2"
    EMP = "員工資料"
    emp_last_row = ws.max_row

    # ---------- 薪資試算 ----------
    ws = wb.create_sheet("薪資試算", 0)
    headers = [
        "員工代碼", "角色", "底薪", "時薪", "上班時數\n(本月)", "加班時數\n(本月)",
        "底薪/工資", "加班費", "勞保投保薪資", "職災投保薪資", "健保投保薪資", "勞退投保薪資",
        "勞保\n(公司)", "職災保險\n(公司)", "健保\n(公司)", "勞退\n(公司)",
        "公司總負擔成本", "勞保(員工自付,僅供參考)", "健保(員工自付,僅供參考)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 34

    N_ROWS = 30
    LAB, OCC, HEA, PEN = "勞保級距表", "職災保險級距表", "健保級距表", "勞退級距表"
    for i in range(N_ROWS):
        r = i + 2
        code_cell = f"A{r}"
        ws.cell(row=r, column=1)  # 員工代碼，輸入
        ws.cell(row=r, column=2, value=f'=IFERROR(VLOOKUP({code_cell},{EMP}!$A$2:$D${emp_last_row},2,FALSE),"")')
        ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP({code_cell},{EMP}!$A$2:$D${emp_last_row},3,FALSE),"")')
        ws.cell(row=r, column=4, value=f'=IFERROR(VLOOKUP({code_cell},{EMP}!$A$2:$D${emp_last_row},4,FALSE),"")')
        # E,F 上班時數/加班時數 = 輸入
        role_cell = f"B{r}"
        base_cell = f"C{r}"
        hourly_cell = f"D{r}"
        hrs_cell = f"E{r}"
        ot_cell = f"F{r}"
        # 兼職且時數還沒填(空白)或填0時要回傳""，不能讓公式算出0元——0元會讓後面查級距表
        # 查到比表上最小值(1,500)還小的數字，VLOOKUP 約略比對找不到，會噴 #N/A。
        # (時數填0通常代表這個月完全沒排班，本來就不該有薪資/保費數字，回傳""正確反映這件事)
        g = (f'=IF({role_cell}="兼職",'
             f'IF(OR({hrs_cell}="",{hrs_cell}=0),"",{hrs_cell}*{hourly_cell}),'
             f'{base_cell})')
        h = f'=IF({role_cell}="兼職",0,{ot_cell}*{p("加班時薪")})'
        ws.cell(row=r, column=7, value=g)
        ws.cell(row=r, column=8, value=h)
        basis = f"G{r}"
        # VLOOKUP範圍故意剛好對齊每張級距表實際資料筆數(不多留空白列)——近似比對(TRUE)
        # 遇到範圍內有空白列會被當成0，打亂由小到大排序的前提，可能悄悄查到錯的級距，
        # 不會跳出明顯的公式錯誤，所以更要注意這個細節。
        # 兼職查勞保/職災表前要先套11,100元下限——官方部分工時特例是月收入未達11,100元
        # 一律用11,100元計算，不是查最接近的細分級距（例如月收入7,840元不會查到8,700元，
        # 是直接用11,100元），這裡用MAX(基準,11100)重現這條規則。
        j = (f'=IF({basis}="","",VLOOKUP(IF({role_cell}="兼職",MAX({basis},11100),MAX({basis},29500)),'
             f'{LAB}!$A$3:$C${LAB_LAST},3,TRUE))')
        k = (f'=IF({basis}="","",VLOOKUP(IF({role_cell}="兼職",MAX({basis},11100),MAX({basis},29500)),'
             f'{OCC}!$A$3:$C${OCC_LAST},3,TRUE))')
        l = f'=IF({basis}="","",VLOOKUP(MAX({basis},29500),{HEA}!$A$3:$C${HEA_LAST},3,TRUE))'
        m = f'=IF({basis}="","",VLOOKUP({basis},{PEN}!$A$3:$C${PEN_LAST},3,TRUE))'
        ws.cell(row=r, column=9, value=j)
        ws.cell(row=r, column=10, value=k)
        ws.cell(row=r, column=11, value=l)
        ws.cell(row=r, column=12, value=m)
        jc, kc, lc, mc = f"I{r}", f"J{r}", f"K{r}", f"L{r}"
        n = (f'=IF({jc}="","",ROUND({jc}*{p("勞保_普通事故費率")}*{p("勞保_公司比例")},0)'
             f'+ROUND({jc}*{p("勞保_就業保險費率")}*{p("勞保_公司比例")},0))')
        o = (f'=IF({kc}="","",ROUND({kc}*{p("職災_行業別費率")},0)'
             f'+ROUND({kc}*{p("職災_上下班費率")},0))')
        q = (f'=IF({lc}="","",ROUND({lc}*{p("健保_費率")}*{p("健保_公司比例")}*{p("健保_公司眷屬平均乘數")},0))')
        s = f'=IF({mc}="","",ROUND({mc}*{p("勞退_費率")},0))'
        ws.cell(row=r, column=13, value=n)
        ws.cell(row=r, column=14, value=o)
        ws.cell(row=r, column=15, value=q)
        ws.cell(row=r, column=16, value=s)
        gcost = f"G{r}"
        hcost = f"H{r}"
        ncost, ocost, qcost, scost = f"M{r}", f"N{r}", f"O{r}", f"P{r}"
        total = f'=IF({gcost}="","",{gcost}+{hcost}+{ncost}+{ocost}+{qcost}+{scost})'
        ws.cell(row=r, column=17, value=total)
        emp_labor = (f'=IF({jc}="","",ROUND({jc}*{p("勞保_普通事故費率")}*{p("勞保_員工比例")},0)'
                     f'+ROUND({jc}*{p("勞保_就業保險費率")}*{p("勞保_員工比例")},0))')
        emp_health = f'=IF({lc}="","",ROUND({lc}*{p("健保_費率")}*{p("健保_員工比例")},0))'
        ws.cell(row=r, column=18, value=emp_labor)
        ws.cell(row=r, column=19, value=emp_health)

    total_row = N_ROWS + 2
    ws.cell(row=total_row, column=1, value="本頁合計").font = BOLD
    for col in (7, 8, 13, 14, 15, 16, 17):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f'=SUM({letter}2:{letter}{total_row - 1})')
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.font = BOLD
        cell.fill = TOTAL_FILL
        cell.border = BORDER

    # 樣式：輸入欄(藍) vs 公式欄(黑) vs 跨分頁引用(綠)
    for r in range(2, N_ROWS + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if c == 1:
                cell.font = BLUE
            elif c in (5, 6):
                cell.font = BLUE
                cell.fill = ASSUMPTION_FILL
            elif c in (2, 3, 4):
                cell.font = GREEN
            elif c == 17:
                cell.font = Font(name=FONT_NAME, bold=True)
            else:
                cell.font = BLACK
    money_cols = [3, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
    for r in range(2, total_row + 1):
        for c in money_cols:
            ws.cell(row=r, column=c).number_format = '#,##0;(#,##0);"-"'
    widths = {"A": 10, "B": 8, "C": 10, "D": 8, "E": 10, "F": 10, "G": 11, "H": 10,
              "I": 12, "J": 12, "K": 12, "L": 12, "M": 9, "N": 9, "O": 9, "P": 9,
              "Q": 13, "R": 15, "S": 15}
    autosize(ws, widths)
    dv_code = DataValidation(type="list", formula1=f"={EMP}!$A$2:$A${emp_last_row}", allow_blank=True)
    ws.add_data_validation(dv_code)
    dv_code.add(f"A2:A{N_ROWS + 1}")
    ws.freeze_panes = "A2"

    # 保護公式欄，只留輸入欄可編輯（無密碼，使用者自己可以隨時解鎖）
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)
    for r in range(2, total_row):
        for c in (2, 3, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19):
            wb["薪資試算"].cell(row=r, column=c).protection = Protection(locked=True)
    wb["薪資試算"].protection.sheet = True

    wb.move_sheet("薪資試算", offset=-len(wb.sheetnames))
    ws_order = ["薪資試算", "員工資料", "參數設定", "勞保級距表", "職災保險級距表", "健保級距表", "勞退級距表", "說明"]
    wb._sheets = [wb[name] for name in ws_order]
    wb.active = 0

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)
    print(f"已產生 {OUT_PATH}")


if __name__ == "__main__":
    main()
