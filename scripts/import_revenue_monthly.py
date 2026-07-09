#!/usr/bin/env python3
"""營收月報匯入程式：把 data/raw 下的營收月報 xlsx 讀進 raw_revenue_monthly。

跨月、跨店欄位名稱不一致，用對照表統一，不能用欄位順序硬抓：
    A 店 202604：消費方式 / 營收 / 營收佔比 / 訂單數 / 單數占比（無門市名稱欄，且是「訂單數」不是「杯數」）
    A 店 202605+：門市名稱 / 單別 / 金額 / 金額百分比 / 杯數
    B 店：門市名稱 / 營收合計 / 折扣 / ... / 現金支出（跟收銀機明細同一套付款方式欄位，但是「整月一列」的加總，
          沒有消費方式拆分）—— 這種檔案只取「營收合計」當單一列存進去，order_type 用固定代號
          MONTHLY_TOTAL 標記，不是真的消費方式類別，之後統計消費方式佔比時要排除這個代號。

用法：
    source .venv/bin/activate
    python3 scripts/import_revenue_monthly.py
"""
import glob
import re
import sqlite3
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "db" / "riva_agent.db"
ENV_PATH = ROOT / ".env"
RAW_GLOB = str(ROOT / "data" / "raw" / "*營收.xlsx")

# 欄位名稱對照表：{統一後欄位: [各月份可能出現的原始欄名]}（先 strip 前後空白再比對）
COLUMN_ALIASES = {
    "store_name": ["門市名稱"],
    "order_type": ["消費方式", "單別"],
    "amount": ["營收", "金額"],
    "pct_of_total": ["營收佔比", "金額百分比"],
    "cup_count": ["杯數"],
    "order_count": ["訂單數"],
    "monthly_total": ["營收合計"],
}

MONTHLY_TOTAL_ORDER_TYPE = "MONTHLY_TOTAL"


def load_store_map():
    store_map = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip()
        if key.startswith("STORE_") and key.endswith("_REAL_NAME") and value:
            store_id = key[len("STORE_"):-len("_REAL_NAME")]
            store_map[value] = store_id
    return store_map


def extract_year_month(filename):
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", filename)
    if not match:
        raise ValueError(f"{filename}：檔名裡找不到 YYYYMM，請確認檔名格式")
    return f"{match.group(1)}-{match.group(2)}"


def build_column_index(header):
    """回傳 {統一後欄位: 原始欄位在 row 中的 index}，找不到的欄位不會出現在結果裡。"""
    stripped_header = [str(h).strip() if h is not None else h for h in header]
    col = {}
    for unified, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in stripped_header:
                col[unified] = stripped_header.index(alias)
                break
    return col


def store_id_from_filename(filename, store_map):
    """檔名裡如果有『X店』標記（例如 B店），直接對到代號 X。"""
    match = re.search(r"([A-Z])店", filename)
    if match and match.group(1) in store_map.values():
        return match.group(1)
    return None


def resolve_store_id(col, row, store_map, filename):
    if "store_name" in col:
        raw_name = row[col["store_name"]]
        store_id = store_map.get(raw_name)
        if store_id is None:
            raise ValueError(
                f"{filename}：門市名稱在 .env 對照表裡找不到（不印出真實店名，避免洩漏），"
                "請確認 STORE_A_REAL_NAME / STORE_B_REAL_NAME"
            )
        return store_id
    # 這份報表沒有門市名稱欄：先看檔名有沒有『X店』標記
    by_filename = store_id_from_filename(filename, store_map)
    if by_filename:
        return by_filename
    # 專案命名慣例：只有非 A 店的檔案才會加上『X店』標記（A 店是最早、唯一店時期留下的檔案，從來不標記）
    # 這是慣例推論、不是直接證據，每次用到都明確印出來，不要讓歸戶悄悄發生
    if "A" in store_map.values():
        print(f"  [慣例判斷] {filename}：無門市名稱欄、檔名也無標記，依慣例歸戶到 A 店，請自行確認是否合理")
        return "A"
    if len(store_map) == 1:
        return next(iter(store_map.values()))
    raise ValueError(
        f"{filename}：這份報表沒有門市名稱欄，檔名裡也沒有『X店』標記，也無法用『無標記=A店』的慣例判斷，"
        "請手動確認後調整程式"
    )


def import_file(path, store_map, conn):
    year_month = extract_year_month(path.name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    col = build_column_index(header)

    if "monthly_total" in col and "order_type" not in col:
        inserted = 0
        for row in data_rows:
            if row[col["monthly_total"]] is None:
                continue
            store_id = resolve_store_id(col, row, store_map, path.name)
            conn.execute(
                """
                INSERT INTO raw_revenue_monthly
                    (store_id, year_month, order_type, amount, pct_of_total,
                     cup_count, order_count, source_file)
                VALUES (?, ?, ?, ?, NULL, NULL, NULL, ?)
                """,
                (store_id, year_month, MONTHLY_TOTAL_ORDER_TYPE, row[col["monthly_total"]], path.name),
            )
            inserted += 1
        return inserted, year_month

    required = {"order_type", "amount"}
    missing = required - col.keys()
    if missing:
        raise ValueError(f"{path.name}：缺少必要欄位 {missing}，欄位對照表可能需要更新")

    inserted = 0
    for row in data_rows:
        if row[col["order_type"]] is None:
            continue
        store_id = resolve_store_id(col, row, store_map, path.name)

        conn.execute(
            """
            INSERT INTO raw_revenue_monthly
                (store_id, year_month, order_type, amount, pct_of_total,
                 cup_count, order_count, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                store_id,
                year_month,
                row[col["order_type"]],
                row[col["amount"]],
                row[col["pct_of_total"]] if "pct_of_total" in col else None,
                row[col["cup_count"]] if "cup_count" in col else None,
                row[col["order_count"]] if "order_count" in col else None,
                path.name,
            ),
        )
        inserted += 1
    return inserted, year_month


def peek_store_and_month(path, store_map):
    """不寫入資料庫，只讀出這份檔案屬於哪個 (store_id, year_month)，供重跑前清資料用。"""
    year_month = extract_year_month(path.name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    col = build_column_index(header)
    key_col = "monthly_total" if ("monthly_total" in col and "order_type" not in col) else "order_type"
    for row in data_rows:
        if key_col in col and row[col[key_col]] is not None:
            return resolve_store_id(col, row, store_map, path.name), year_month
    raise ValueError(f"{path.name}：整份檔案沒有任何有效資料列，無法判斷店別")


def main():
    store_map = load_store_map()
    if not store_map:
        raise SystemExit(".env 裡的 STORE_A_REAL_NAME / STORE_B_REAL_NAME 都是空的，請先填店名對照")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    files = sorted(Path(p) for p in glob.glob(RAW_GLOB))
    if not files:
        raise SystemExit(f"找不到符合 {RAW_GLOB} 的檔案")

    # 重跑時先清掉「這次要匯入的店+月份」的舊資料，用 (store_id, year_month) 而不是檔名比對，
    # 避免訂正檔換了檔名後，舊檔名底下的舊資料變孤兒、永遠留在表裡疊加。
    periods_to_refresh = {peek_store_and_month(path, store_map) for path in files}
    conn.executemany(
        "DELETE FROM raw_revenue_monthly WHERE store_id = ? AND year_month = ?",
        list(periods_to_refresh),
    )

    total = 0
    for path in files:
        count, year_month = import_file(path, store_map, conn)
        print(f"{path.name} ({year_month}): 匯入 {count} 筆")
        total += count

    conn.commit()
    conn.close()
    print(f"完成，總計匯入 {total} 筆到 raw_revenue_monthly")


if __name__ == "__main__":
    main()
